#!/usr/bin/python3 
import argparse
import numpy as np
import sys
import time

from openpyxl import load_workbook

users = {}
titlesUsers = {}
titlesDiagnoses = {}

def parseSheetOne(sheet):
    global users
    global titlesUsers

    # Unwrap into values
    unwrapped_values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    titlesUsers = unwrapped_values[0]

    # Convert to dictionary
    users = {row[0]:row[1:] for row in unwrapped_values}
    del users['EnterpriseID']

def parseSheetTwo(sheet):
    global users
    global titlesDiagnoses

    # Unwrap into values
    encounters = [[cell.value for cell in row] for row in sheet.iter_rows()]
    titlesDiagnoses = encounters[0]
    del encounters[0]

    # Add to existing dictionary
    for encounter in encounters:
        user = encounter[0]

        if user in users:
            users[user] = users[user] + encounter[1:]
        else:
            sys.exit('Encountered User in Sheet2 without corresponding Enterprise ID in Sheet 1.')

def addNewSheet(workbook):
    global users
    global titlesUsers
    global titlesDiagnoses

    ws = workbook.create_sheet('Output')

    for key, value in users.items():
        row = [key] + value
        ws.append(row)

    # Generate and write title to start. This is inefficient, fix later
    encounter_count = (ws.max_column - len(titlesUsers)) / (len(titlesDiagnoses) - 1)
    del titlesDiagnoses[0]
    expanded_titles = []
    for i in range(int(encounter_count)):
        expanded_titles += [title + f' {i}' for title in titlesDiagnoses]
    titles = titlesUsers + expanded_titles

    ws.insert_rows(0)
    for i in range(len(titles)):
        ws.cell(column=i+1, row=1, value=titles[i])

    result_document = f'Results_{str(int(time.time()))}.xlsx'
    workbook.save(result_document)
    print(f'Your result is saved in {result_document}')


def main():
    parser = argparse.ArgumentParser(description='Convert Pavan\'s documents.')
    parser.add_argument('--source', dest='source', required=True,
                    help='The source document.')

    args = parser.parse_args()

    wb = load_workbook(args.source, data_only=True)

    # Create model of users and append Sheet2's data
    parseSheetOne(wb['Sheet1'])  
    parseSheetTwo(wb['Sheet2'])
    addNewSheet(wb)

    # Output and save new sheet

if __name__ == '__main__':
    main()
